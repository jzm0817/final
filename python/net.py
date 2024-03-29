
import path
import par
from plotcm import plot_confusion_matrix
from openpyxl import Workbook

import numpy as np
import torch
import torch.nn as nn
import torch.nn.functional as file
import torchvision
from collections import OrderedDict
from sklearn.metrics import confusion_matrix
import matplotlib.pyplot as plt


class trainpar():
    def __init__(self, batch_size, learning_rate, epoch,  ann, pic_size, pic_list, pic_enhance_list):
        self.pic_size = pic_size
        self.pic_list = pic_list
        self.pic_enhance_list = pic_enhance_list
        self.bs = batch_size
        self.lr = learning_rate
        self.epoch = epoch
        self.ann = ann


class residual_block(nn.Module):
    def __init__(self, ch_in, ch_out, stride=1, downsample=None):
        super().__init__()
        self.conv1 = nn.Conv2d(ch_in, ch_out, kernel_size=3, stride=stride, padding=1, bias=False)
        self.bn1 = nn.BatchNorm2d(ch_out)
        self.relu = nn.ReLU(inplace=True)
        
        self.conv2 = nn.Conv2d(ch_out, ch_out, kernel_size=3, stride=1, padding=1, bias=False)
        self.bn2 = nn.BatchNorm2d(ch_out)
        self.downsample = downsample

    
    def forward(self, x):
        residual = x
        
        out = self.conv1(x)
        out = self.bn1(out)
        out = self.relu(out)
        out = self.conv2(out)
        out = self.bn2(out)
        
        if self.downsample:
            residual = self.downsample(x)
        
        out += residual
        out = self.relu(out)

        return out

class resnet(nn.Module):
    def __init__(self, block, layers, num_classes=4):
        super().__init__()
        self.ch_in = 16
        self.conv = nn.Conv2d(3, 16, kernel_size=3)
        self.bn = nn.BatchNorm2d(16)
        self.relu = nn.ReLU(inplace=True)
        self.layer1 = self.make_layer(block, 16, layers[0])
        self.layer2 = self.make_layer(block, 32, layers[1], 2)
        self.layer3 = self.make_layer(block, 64, layers[2], 2)
        self.avg_pool = nn.AvgPool2d(8)
        self.fc = nn.Linear(64 * 6 * 6, num_classes)
    
    def make_layer(self, block, ch_out, blocks, stride=1):
        downsample = None
        if (stride != 1) or (self.ch_in != ch_out):
            downsample = nn.Sequential(
            nn.Conv2d(self.ch_in, ch_out, kernel_size=3, stride=stride, padding=1, bias=False),
            nn.BatchNorm2d(ch_out)
                                     )
        layers = []
        layers.append(block(self.ch_in, ch_out, stride, downsample))
        self.ch_in = ch_out
        for i in range(1, blocks):
            layers.append(block(ch_out, ch_out))
        return nn.Sequential(*layers)# add all of the residual block
            
    
    def forward(self,x):
        out = self.conv(x) 
        out = self.bn(out) 
        out = self.relu(out) 
        out = self.layer1(out) 
        out = self.layer2(out) 
        out = self.layer3(out) 
        out = self.avg_pool(out) 
        out = out.view(out.size(0), -1) 
        out = self.fc(out) 
        
        return out 

class neuralnetwork(nn.Module):

    def __init__(self, nn_list):
        super().__init__()
        self.nn_list = nn_list
        self.ann = nn.Sequential(self.nn_list)
    
    def forward(self, x):
        return self.ann(x)



def train(model, data_set_training, optimizer, loss_fn, epoch, device):
    model.train()

    loss_total = 0

    for _, data in enumerate(data_set_training):
        if isinstance(data, list):
            image = data[0].type(torch.FloatTensor).to(device)
            label = data[1].to(device)
        elif isinstance(data, dict):
            image = data['image'].type(torch.FloatTensor).to(device)
            label = data['label'].to(device)
        else:
            print(type(data))
            raise TypeError
        
        optimizer.zero_grad()
        # print(f'image.shape:{image.shape}')
        output = model(image)

        loss = loss_fn(output, label)
        loss_total+=loss.item()

        loss.backward()
        optimizer.step()
    
    if (epoch + 1) % 10 == 0:
        print(f'{round(loss_total,2)} in epoch {epoch + 1}')
    return loss_total


def test(model, data_set_test, device, bs, trained_name, mul_flag):
    model.eval()

    correct = 0
    real_label = []
    pred_label = []
    bs = 0

    for ii, data in enumerate(data_set_test):

        if isinstance(data, list):
            image = data[0].type(torch.FloatTensor).to(device)
            label = data[1].to(device)
        elif isinstance(data, dict):
            image = data['image'].type(torch.FloatTensor).to(device)
            label = data['label'].to(device)
        else:
            print(type(data))
            raise TypeError
        
        with torch.no_grad():
            output = model(image)
            pred = nn.Softmax(dim=1)(output)
        
        real_label = np.append(real_label, label.cpu().numpy())
        pred_label = np.append(pred_label, pred.argmax(1).cpu().numpy())

        if ii == 0:
            bs = len(label.cpu().numpy().tolist())
            # print(f'label:{len(real_label)}')
            # print(f'label.numpy():{real_label}')
            # print(f'pred.argmax(1):{pred_label}')

        correct += (pred.argmax(1) == label).type(torch.float).sum().item()


    real_label = np.array(real_label)
    pred_label = np.array(pred_label)

    if mul_flag:
        wb = Workbook()
        ws = wb.active
        # ws['A1'] = 'real_label'
        # ws['B1'] = 'pred_label'
        row1 = real_label.tolist()
        row2 = pred_label.tolist()
        ws.append(row1)
        ws.append(row2)
        row3 = []
        row4 = []
        mem_num = 7
        for i in range(0, len(row1[::mem_num ])):
            vec = row2[(i) * mem_num:(i+1)* mem_num]
            vec_set = set(vec)
            vec_tmp = np.bincount(vec)
            row3.append(np.argmax(vec_tmp))  
            if len(vec_set) == mem_num:
                row4.append(1)
            else:
                row4.append(0)
        ws.append(row3)
        ws.append(row4)
        # print(len(row1[::3]))
        # print(pred_label)
        # print(real_label)
        wb.save('D:/workspace/art/net/' + trained_name + '.xlsx')
        # print(real_label.flatten())
        real_label_ = row1[::mem_num ]
        
        cnt_err = 0
        cnt_tr = 0
        real_label_mod = []
        pred_label_mod = []

        for i in range(0, len(real_label_)):
            if row4[i] == 1:
                cnt_err += 1
            else:
                cnt_tr += 1
                real_label_mod.append(real_label_[i])
                pred_label_mod.append(row3[i])
        

        print(f'cnt_err:{cnt_err}')
        print(f'cnt_tr:{cnt_tr}') 
        # print(real_label_mod)
        # print(pred_label_mod)
        stacked1 = torch.stack((torch.tensor(real_label_mod, dtype=torch.int64), torch.tensor(pred_label_mod, dtype=torch.int64)), dim=1)
        names1 = list(par.data_type_dict.values())
        cm1 = torch.zeros(len(par.data_type_dict), len(par.data_type_dict), dtype=torch.int64)
        for p in stacked1:
            tl, pl = p.tolist()
            cm1[tl, pl] = cm1[tl, pl] + 1
        # print(cm1)
        plot_confusion_matrix(cm1, names1, normalize=False)
        pres = "mul_"
        plt.savefig(f"{path.trainednet_path}/{pres}nn{trained_name}.png")


    stacked = torch.stack((torch.tensor(real_label.flatten(), dtype=torch.int64), torch.tensor(pred_label.flatten(), dtype=torch.int64)), dim=1)
    # print(stacked[0:7, :])
    names = list(par.data_type_dict.values())
    # print(names)
   
    cm = torch.zeros(len(par.data_type_dict), len(par.data_type_dict), dtype=torch.int64)
    for p in stacked:
        tl, pl = p.tolist()
        cm[tl, pl] = cm[tl, pl] + 1
    # print(cm)
    # cm = torch(np.array([[98, 0, 0, 102], [0, 194, 3, 3], [0, 71, 128, 1], [1, 0, 0, 199]]) , dtype=torch.int64)
    
    plot_confusion_matrix(cm, names, normalize=True)
    pres = "cm_"
    output_str = "_normal"
    plt.savefig(f"{path.trainednet_path}/{pres}nn{trained_name + output_str}.png")
    plt.figure()
    plot_confusion_matrix(cm, names)
    plt.savefig(f"{path.trainednet_path}/{pres}nn{trained_name}.png")
    # plt.show()
    # print(f'stacked.shape:{stacked.shape}')
    # print(f'len(real_label):{len(real_label)}') 
    # print(f'len(real_label):{len(real_label)}') 
    # print(len(data_set_test))
    print(f'accuracy = {correct}/{len(data_set_test) * bs} = {correct/(len(data_set_test) * bs)}')
    # print(f'accuracy = {correct}/{len(data_set_test) * bs} = {correct/len(data_set_test) / bs}')




